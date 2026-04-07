"""
Baseline
────────
A production-ready Flask starter kit with authentication, admin dashboard,
account management, and GDPR compliance. Fork this repo and build your
app-specific features on top.

Architecture:
  - Auth: login, register, password reset, email verification
  - Account: profile, password, avatar, GDPR export/delete
  - Admin: user management, impersonation, platform stats
  - Example CRUD: notes (delete when you build your own features)
"""

import csv
import io
import json
import os
import re
import sqlite3
import time
from datetime import datetime, timezone
from functools import wraps
from pathlib import Path

from dotenv import load_dotenv
from flask import (
    Flask, render_template, request,
    redirect, url_for, session, flash, g, abort, send_from_directory, jsonify
)
from flask_wtf.csrf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from markupsafe import escape, Markup
import markdown as md_lib
import bleach
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import urllib.request
import urllib.error

# ── Config ────────────────────────────────────────────────────────

load_dotenv()

APP_DIR = Path(__file__).parent
DB_PATH = APP_DIR / "data" / "app.db"

UPLOAD_DIR = APP_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
ALLOWED_EXTENSIONS = {"pdf", "doc", "docx", "ppt", "pptx", "xls", "xlsx", "txt", "csv", "zip",
                      "png", "jpg", "jpeg", "gif", "svg", "mp4", "mov", "webm"}
MAX_UPLOAD_MB = 50

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-me-in-production")
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_MB * 1024 * 1024
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("FLASK_ENV") == "production"
app.config["PERMANENT_SESSION_LIFETIME"] = 86400 * 7  # 7 days
app.config["WTF_CSRF_TIME_LIMIT"] = 3600  # 1 hour CSRF token validity
DEFAULT_PORT = int(os.environ.get("PORT", 5000))

# App metadata — update these for your project
APP_NAME = os.environ.get("APP_NAME", "Baseline")
APP_SUPPORT_EMAIL = os.environ.get("SUPPORT_EMAIL", "support@example.com")

# Refuse to boot with default secret key in production
if os.environ.get("FLASK_ENV") == "production" and app.secret_key == "change-me-in-production":
    raise RuntimeError("SECRET_KEY must be set in production. Generate one with: python3 -c \"import secrets; print(secrets.token_hex(32))\"")

# CSRF protection
csrf = CSRFProtect(app)

# Rate limiting
limiter = Limiter(get_remote_address, app=app, default_limits=["200 per minute"],
                  storage_uri="memory://")

# Security headers
@app.after_request
def set_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    if request.is_secure:
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response

(APP_DIR / "data").mkdir(exist_ok=True)

# ── Markdown Rendering ────────────────────────────────────────────

ALLOWED_TAGS = [
    "p", "br", "strong", "em", "a", "ul", "ol", "li", "code", "pre",
    "blockquote", "h1", "h2", "h3", "h4", "img", "hr", "del", "table",
    "thead", "tbody", "tr", "th", "td",
]
ALLOWED_ATTRS = {
    "a": ["href", "title", "rel"],
    "img": ["src", "alt", "title"],
}

def render_markdown(text):
    """Convert Markdown to sanitised HTML."""
    if not text:
        return ""
    raw_html = md_lib.markdown(text, extensions=["fenced_code", "tables", "nl2br"])
    clean = bleach.clean(raw_html, tags=ALLOWED_TAGS, attributes=ALLOWED_ATTRS)
    clean = clean.replace("<a ", '<a target="_blank" rel="noopener" ')
    return Markup(clean)

@app.template_filter("markdown")
def markdown_filter(text):
    return render_markdown(text)

def strip_markdown(text):
    """Remove markdown syntax for plain text previews."""
    if not text:
        return ""
    t = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
    t = re.sub(r'\*(.+?)\*', r'\1', t)
    t = re.sub(r'__(.+?)__', r'\1', t)
    t = re.sub(r'_(.+?)_', r'\1', t)
    t = re.sub(r'#{1,6}\s*', '', t)
    t = re.sub(r'^\s*[-*+]\s+', '', t, flags=re.MULTILINE)
    t = re.sub(r'^\s*\d+\.\s+', '', t, flags=re.MULTILINE)
    t = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', t)
    t = re.sub(r'`{1,3}[^`]*`{1,3}', '', t)
    t = re.sub(r'^>\s*', '', t, flags=re.MULTILINE)
    t = re.sub(r'---+', '', t)
    t = re.sub(r'\n{2,}', ' ', t)
    return t.strip()

@app.template_filter("timeago")
def timeago_filter(dt_str):
    """Convert ISO datetime string to relative time like '3h ago'."""
    if not dt_str:
        return ""
    try:
        dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
    except (ValueError, AttributeError):
        return dt_str[:10] if dt_str else ""
    now = datetime.now(timezone.utc)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    diff = now - dt
    seconds = int(diff.total_seconds())
    if seconds < 60:
        return "just now"
    elif seconds < 3600:
        m = seconds // 60
        return f"{m}m ago"
    elif seconds < 86400:
        h = seconds // 3600
        return f"{h}h ago"
    elif seconds < 604800:
        d = seconds // 86400
        return f"{d}d ago"
    elif seconds < 2592000:
        w = seconds // 604800
        return f"{w}w ago"
    else:
        return dt_str[:10]

@app.template_filter("strip_markdown")
def strip_markdown_filter(text):
    return strip_markdown(text)

# ── Email (Resend) ────────────────────────────────────────────────

RESEND_API_KEY = os.environ.get("RESEND_API_KEY", "")
EMAIL_FROM = os.environ.get("EMAIL_FROM", f"{APP_NAME} <noreply@example.com>")

def send_email(to, subject, html_body):
    """Send email via Resend API. Returns True on success."""
    if not RESEND_API_KEY:
        print(f"[EMAIL SKIPPED — no API key] To: {to}, Subject: {subject}")
        return False
    payload = json.dumps({
        "from": EMAIL_FROM,
        "to": [to],
        "subject": subject,
        "html": html_body
    }).encode()
    req = urllib.request.Request(
        "https://api.resend.com/emails",
        data=payload,
        headers={"Authorization": f"Bearer {RESEND_API_KEY}", "Content-Type": "application/json"},
        method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            return resp.status == 200
    except urllib.error.URLError as e:
        print(f"[EMAIL ERROR] {e}")
        return False

# ── Token Generation ──────────────────────────────────────────────

from itsdangerous import URLSafeTimedSerializer
_serializer = URLSafeTimedSerializer(app.secret_key)

def generate_token(data, salt="default"):
    return _serializer.dumps(data, salt=salt)

def verify_token(token, salt="default", max_age=3600):
    try:
        return _serializer.loads(token, salt=salt, max_age=max_age)
    except Exception:
        return None

# ── Database ──────────────────────────────────────────────────────

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(str(DB_PATH))
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db

@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db:
        db.close()

def init_db():
    db = sqlite3.connect(str(DB_PATH))
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA foreign_keys=ON")
    db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            name          TEXT NOT NULL,
            email         TEXT NOT NULL UNIQUE COLLATE NOCASE,
            password_hash TEXT NOT NULL,
            is_admin      INTEGER NOT NULL DEFAULT 0,
            email_verified INTEGER NOT NULL DEFAULT 0,
            bio           TEXT DEFAULT '',
            location      TEXT DEFAULT '',
            website       TEXT DEFAULT '',
            avatar_path   TEXT DEFAULT '',
            created       TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS notes (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id       INTEGER NOT NULL,
            title         TEXT NOT NULL,
            body          TEXT DEFAULT '',
            created       TEXT NOT NULL,
            updated       TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS gantt_groups (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id    INTEGER NOT NULL,
            name       TEXT NOT NULL,
            created    TEXT NOT NULL,
            updated    TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS gantt_charts (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER NOT NULL,
            name        TEXT NOT NULL,
            weeks_json  TEXT NOT NULL DEFAULT '[]',
            group_id    INTEGER DEFAULT NULL,
            logo_path   TEXT DEFAULT '',
            brand_color TEXT DEFAULT '#2a5a8a',
            created     TEXT NOT NULL,
            updated     TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
            FOREIGN KEY (group_id) REFERENCES gantt_groups(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS gantt_rows (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            chart_id   INTEGER NOT NULL,
            row_type   TEXT NOT NULL DEFAULT 'task',
            action     TEXT NOT NULL DEFAULT '',
            owner      TEXT DEFAULT '',
            hours      REAL DEFAULT 0,
            objective  TEXT DEFAULT '',
            kpi        TEXT DEFAULT '',
            status     TEXT DEFAULT 'not_started',
            start_week TEXT DEFAULT '',
            end_week   TEXT DEFAULT '',
            notes_json TEXT DEFAULT '{}',
            sort_order INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY (chart_id) REFERENCES gantt_charts(id) ON DELETE CASCADE
        );
    """)

    # Run migrations
    run_migrations(db)

    # Create default platform admin if no users exist
    row = db.execute("SELECT COUNT(*) FROM users").fetchone()
    if row[0] == 0:
        db.execute(
            "INSERT INTO users (name, email, password_hash, is_admin, email_verified, created) VALUES (?, ?, ?, ?, ?, ?)",
            ("Admin", "admin@example.com", generate_password_hash("changeme"), 1, 1, datetime.now(timezone.utc).isoformat())
        )
        print(f"Default admin created — email: admin@example.com  password: changeme")

    db.commit()
    db.close()

def run_migrations(db):
    """Run numbered SQL migration files from migrations/ directory."""
    migrations_dir = APP_DIR / "migrations"
    if not migrations_dir.exists():
        return
    db.execute("""
        CREATE TABLE IF NOT EXISTS _migrations (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            applied TEXT NOT NULL
        )
    """)
    applied = {row[0] for row in db.execute("SELECT name FROM _migrations").fetchall()}
    migration_files = sorted(migrations_dir.glob("*.sql"))
    for f in migration_files:
        if f.name not in applied:
            print(f"[MIGRATE] Applying {f.name}")
            db.executescript(f.read_text())
            db.execute("INSERT INTO _migrations (name, applied) VALUES (?, ?)",
                       (f.name, datetime.now(timezone.utc).isoformat()))
    db.commit()

# ── Helpers ───────────────────────────────────────────────────────

def get_user_by_id(uid):
    return get_db().execute("SELECT * FROM users WHERE id = ?", (uid,)).fetchone()

def get_user_by_email(email):
    return get_db().execute("SELECT * FROM users WHERE email = ? COLLATE NOCASE", (email,)).fetchone()

def current_user():
    uid = session.get("user_id")
    if uid:
        return get_user_by_id(uid)
    return None

def login_user(user):
    """Set session for user with session regeneration to prevent fixation."""
    session.clear()
    session["user_id"] = user["id"]
    session["user_name"] = user["name"]
    session.permanent = True

def slugify(text):
    slug = re.sub(r'[^a-z0-9]+', '-', text.lower()).strip('-')
    return slug[:60]

def avatar_html(user, size=22):
    """Generate avatar HTML — photo if available, initial letter if not."""
    if user and user["avatar_path"]:
        return Markup(f'<img src="/uploads/{user["avatar_path"]}" class="avatar-img" style="width:{size}px;height:{size}px;" />')
    name = user["name"] if user else "?"
    return Markup(f'<span class="avatar" style="width:{size}px;height:{size}px;font-size:{max(10, size//2)}px;">{name[0].upper()}</span>')

def paginate(query, params, page, per_page=20):
    """Add LIMIT/OFFSET to a query and return (items, total, pages)."""
    db = get_db()
    count_q = f"SELECT COUNT(*) FROM ({query})"
    total = db.execute(count_q, params).fetchone()[0]
    pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, pages))
    offset = (page - 1) * per_page
    items = db.execute(f"{query} LIMIT ? OFFSET ?", (*params, per_page, offset)).fetchall()
    return items, total, pages, page

def search_like(term):
    """Escape a search term for safe LIKE queries. Returns (pattern, param)."""
    escaped = term.replace("%", "\\%").replace("_", "\\_")
    return f"%{escaped}%"

@app.context_processor
def inject_globals():
    return dict(
        current_user=current_user(),
        avatar_html=avatar_html,
        app_name=APP_NAME,
        support_email=APP_SUPPORT_EMAIL,
    )

# ── Auth Decorators ───────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated

def platform_admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user = current_user()
        if not user or not user["is_admin"]:
            flash("Admin access required.", "error")
            return redirect("/dashboard")
        return f(*args, **kwargs)
    return decorated

# ── Auth Routes ───────────────────────────────────────────────────

@app.route("/")
def landing():
    if session.get("user_id"):
        return redirect("/gantt")
    return render_template("landing.html")

@app.route("/login", methods=["GET", "POST"])
@limiter.limit("10 per minute", methods=["POST"])
def login_page():
    if session.get("user_id"):
        return redirect("/dashboard")
    err = None
    form_ts = str(int(time.time()))
    if request.method == "POST":
        # Honeypot — bots fill hidden field
        if request.form.get("website_url", ""):
            err = "Invalid email or password."
            return render_template("login.html", err=err, form_ts=form_ts)
        # Time trap — reject instant submissions
        ts = request.form.get("_ts", "0")
        try:
            if time.time() - int(ts) < 1.5:
                err = "Invalid email or password."
                return render_template("login.html", err=err, form_ts=form_ts)
        except (ValueError, TypeError):
            pass
        user = get_user_by_email(request.form.get("email", ""))
        if user and check_password_hash(user["password_hash"], request.form.get("password", "")):
            login_user(user)
            return redirect("/dashboard")
        err = "Invalid email or password."
    return render_template("login.html", err=err, form_ts=form_ts)

@app.route("/register", methods=["GET", "POST"])
@limiter.limit("5 per minute", methods=["POST"])
def register_page():
    if session.get("user_id"):
        return redirect("/dashboard")
    err = None
    name = ""
    email = ""
    form_ts = str(int(time.time()))
    if request.method == "POST":
        if request.form.get("website_url", ""):
            err = "Registration failed."
            return render_template("register.html", err=err, name="", email="", form_ts=form_ts)
        ts = request.form.get("_ts", "0")
        try:
            if time.time() - int(ts) < 2:
                err = "Please slow down and try again."
                return render_template("register.html", err=err, name="", email="", form_ts=form_ts)
        except (ValueError, TypeError):
            pass
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        password_confirm = request.form.get("password_confirm", "")
        if not name or not email or not password:
            err = "All fields are required."
        elif len(password) < 8:
            err = "Password must be at least 8 characters."
        elif password != password_confirm:
            err = "Passwords do not match."
        elif get_user_by_email(email):
            err = "An account with that email already exists."
        else:
            db = get_db()
            db.execute(
                "INSERT INTO users (name, email, password_hash, created) VALUES (?, ?, ?, ?)",
                (name, email, generate_password_hash(password), datetime.now(timezone.utc).isoformat())
            )
            db.commit()
            user = get_user_by_email(email)
            login_user(user)
            token = generate_token(user["id"], salt="email-verify")
            verify_url = request.host_url.rstrip("/") + f"/verify-email/{token}"
            send_email(
                to=email,
                subject=f"Verify your email - {APP_NAME}",
                html_body=f"""
                <div style="font-family:sans-serif;max-width:480px;margin:0 auto;padding:32px;">
                    <h2 style="color:#333;">Welcome to {APP_NAME}!</h2>
                    <p style="color:#666;">Hi {name},</p>
                    <p style="color:#666;">Please verify your email address to get full access.</p>
                    <p style="margin:24px 0;">
                        <a href="{verify_url}" style="background:#000;color:#fff;padding:12px 24px;border-radius:4px;text-decoration:none;font-weight:500;">Verify Email</a>
                    </p>
                </div>
                """
            )
            flash("Welcome! Check your email to verify your account.", "success")
            return redirect("/dashboard")
    return render_template("register.html", err=err, name=name, email=email, form_ts=form_ts)

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

# ── Password Reset ────────────────────────────────────────────────

@app.route("/forgot-password", methods=["GET", "POST"])
@limiter.limit("3 per minute", methods=["POST"])
def forgot_password():
    sent = False
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        user = get_user_by_email(email)
        if user:
            token = generate_token(user["id"], salt="password-reset")
            reset_url = request.host_url.rstrip("/") + f"/reset-password/{token}"
            send_email(
                to=user["email"],
                subject="Reset your password",
                html_body=f"""
                <div style="font-family:sans-serif;max-width:480px;margin:0 auto;padding:32px;">
                    <h2 style="color:#333;">Reset your password</h2>
                    <p style="color:#666;">Hi {user['name']},</p>
                    <p style="color:#666;">Click the button below to reset your password. This link expires in 1 hour.</p>
                    <p style="margin:24px 0;">
                        <a href="{reset_url}" style="background:#000;color:#fff;padding:12px 24px;border-radius:4px;text-decoration:none;font-weight:500;">Reset Password</a>
                    </p>
                    <p style="color:#999;font-size:13px;">If you didn't request this, you can safely ignore this email.</p>
                </div>
                """
            )
        sent = True
    return render_template("forgot_password.html", sent=sent)

@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    user_id = verify_token(token, salt="password-reset", max_age=3600)
    if not user_id:
        flash("This reset link has expired or is invalid.", "error")
        return redirect("/forgot-password")
    user = get_user_by_id(user_id)
    if not user:
        flash("User not found.", "error")
        return redirect("/forgot-password")
    err = None
    if request.method == "POST":
        password = request.form.get("password", "")
        password_confirm = request.form.get("password_confirm", "")
        if len(password) < 8:
            err = "Password must be at least 8 characters."
        elif password != password_confirm:
            err = "Passwords do not match."
        else:
            db = get_db()
            db.execute("UPDATE users SET password_hash = ? WHERE id = ?",
                       (generate_password_hash(password), user_id))
            db.commit()
            flash("Password updated! You can now log in.", "success")
            return redirect("/login")
    return render_template("reset_password.html", err=err, token=token)

# ── Email Verification ────────────────────────────────────────────

@app.route("/verify-email/<token>")
def verify_email(token):
    user_id = verify_token(token, salt="email-verify", max_age=86400)
    if not user_id:
        flash("This verification link has expired or is invalid.", "error")
        return redirect("/login")
    db = get_db()
    db.execute("UPDATE users SET email_verified = 1 WHERE id = ?", (user_id,))
    db.commit()
    flash("Email verified!", "success")
    if session.get("user_id"):
        return redirect("/dashboard")
    return redirect("/login")

@app.route("/resend-verification", methods=["POST"])
@login_required
@limiter.limit("2 per minute")
def resend_verification():
    user = get_user_by_id(session["user_id"])
    if user and not user["email_verified"]:
        token = generate_token(user["id"], salt="email-verify")
        verify_url = request.host_url.rstrip("/") + f"/verify-email/{token}"
        send_email(
            to=user["email"],
            subject="Verify your email",
            html_body=f"""
            <div style="font-family:sans-serif;max-width:480px;margin:0 auto;padding:32px;">
                <h2 style="color:#333;">Verify your email</h2>
                <p style="color:#666;">Hi {user['name']},</p>
                <p style="color:#666;">Click the button below to verify your email address.</p>
                <p style="margin:24px 0;">
                    <a href="{verify_url}" style="background:#000;color:#fff;padding:12px 24px;border-radius:4px;text-decoration:none;font-weight:500;">Verify Email</a>
                </p>
            </div>
            """
        )
        flash("Verification email sent! Check your inbox.", "success")
    return redirect("/dashboard")

# ── Dashboard ─────────────────────────────────────────────────────

@app.route("/dashboard")
@login_required
def dashboard():
    db = get_db()
    notes = db.execute(
        "SELECT * FROM notes WHERE user_id = ? ORDER BY updated DESC",
        (session["user_id"],)
    ).fetchall()
    return render_template("dashboard.html", notes=notes)

# ── Account ───────────────────────────────────────────────────────

@app.route("/account")
@login_required
def account_page():
    user = get_user_by_id(session["user_id"])
    return render_template("account.html", user=user)

@app.route("/account/profile", methods=["POST"])
@login_required
def account_profile():
    db = get_db()
    name = request.form.get("name", "").strip()
    email = request.form.get("email", "").strip().lower()
    bio = request.form.get("bio", "").strip()
    location = request.form.get("location", "").strip()
    website = request.form.get("website", "").strip()
    existing = get_user_by_email(email)
    if existing and existing["id"] != session["user_id"]:
        flash("Email already in use.", "error")
        return redirect("/account")
    db.execute(
        "UPDATE users SET name=?, email=?, bio=?, location=?, website=? WHERE id=?",
        (name, email, bio, location, website, session["user_id"])
    )
    db.commit()
    session["user_name"] = name
    flash("Profile updated.", "success")
    return redirect("/account")

@app.route("/account/password", methods=["POST"])
@login_required
def account_password():
    db = get_db()
    user = get_user_by_id(session["user_id"])
    if not check_password_hash(user["password_hash"], request.form.get("current", "")):
        flash("Current password incorrect.", "error")
        return redirect("/account")
    new_pw = request.form.get("new", "")
    if len(new_pw) < 8:
        flash("Password must be at least 8 characters.", "error")
        return redirect("/account")
    db.execute(
        "UPDATE users SET password_hash=? WHERE id=?",
        (generate_password_hash(new_pw), session["user_id"])
    )
    db.commit()
    flash("Password updated.", "success")
    return redirect("/account")

@app.route("/account/avatar", methods=["POST"])
@login_required
def account_avatar():
    file = request.files.get("avatar")
    if not file or not file.filename:
        flash("No file selected.", "error")
        return redirect("/account")
    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("jpg", "jpeg", "png", "webp", "gif"):
        flash("Please upload a JPG, PNG, or WebP image.", "error")
        return redirect("/account")
    avatar_dir = UPLOAD_DIR / "avatars"
    avatar_dir.mkdir(exist_ok=True)
    avatar_name = f"avatar_{session['user_id']}.{ext}"
    file.save(str(avatar_dir / avatar_name))
    db = get_db()
    db.execute("UPDATE users SET avatar_path = ? WHERE id = ?",
               (f"avatars/{avatar_name}", session["user_id"]))
    db.commit()
    flash("Profile photo updated!", "success")
    return redirect("/account")

@app.route("/account/export")
@login_required
def account_export():
    """GDPR data export — download all your data as JSON."""
    db = get_db()
    uid = session["user_id"]
    user = get_user_by_id(uid)
    data = {
        "account": {
            "id": user["id"], "name": user["name"], "email": user["email"],
            "bio": user["bio"], "location": user["location"], "website": user["website"],
            "created": user["created"]
        },
        "notes": [dict(r) for r in db.execute(
            "SELECT id, title, body, created, updated FROM notes WHERE user_id = ?", (uid,)).fetchall()],
        "exported_at": datetime.now(timezone.utc).isoformat()
    }
    return json.dumps(data, indent=2), 200, {
        "Content-Type": "application/json",
        "Content-Disposition": f"attachment; filename=my-data-{uid}.json"
    }

@app.route("/account/delete", methods=["POST"])
@login_required
def account_delete():
    """GDPR right to deletion — delete account and all associated data."""
    uid = session["user_id"]
    db = get_db()
    db.execute("DELETE FROM users WHERE id = ?", (uid,))
    db.commit()
    session.clear()
    flash("Your account and all data have been permanently deleted.", "success")
    return redirect("/")

# ── File Uploads ──────────────────────────────────────────────────

@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    return send_from_directory(str(UPLOAD_DIR), filename)

# ══════════════════════════════════════════════════════════════════
#  EXAMPLE CRUD — Notes (delete this section when building your app)
# ══════════════════════════════════════════════════════════════════

@app.route("/notes/new", methods=["GET", "POST"])
@login_required
def new_note():
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        body = request.form.get("body", "").strip()
        if not title:
            flash("Title is required.", "error")
            return redirect("/notes/new")
        db = get_db()
        now = datetime.now(timezone.utc).isoformat()
        db.execute(
            "INSERT INTO notes (user_id, title, body, created, updated) VALUES (?, ?, ?, ?, ?)",
            (session["user_id"], title, body, now, now)
        )
        db.commit()
        flash("Note created.", "success")
        return redirect("/dashboard")
    return render_template("notes/new.html")

@app.route("/notes/<int:nid>")
@login_required
def view_note(nid):
    db = get_db()
    note = db.execute("SELECT * FROM notes WHERE id = ? AND user_id = ?",
                      (nid, session["user_id"])).fetchone()
    if not note:
        abort(404)
    return render_template("notes/view.html", note=note)

@app.route("/notes/<int:nid>/edit", methods=["GET", "POST"])
@login_required
def edit_note(nid):
    db = get_db()
    note = db.execute("SELECT * FROM notes WHERE id = ? AND user_id = ?",
                      (nid, session["user_id"])).fetchone()
    if not note:
        abort(404)
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        body = request.form.get("body", "").strip()
        if not title:
            flash("Title is required.", "error")
            return redirect(f"/notes/{nid}/edit")
        db.execute(
            "UPDATE notes SET title=?, body=?, updated=? WHERE id=?",
            (title, body, datetime.now(timezone.utc).isoformat(), nid)
        )
        db.commit()
        flash("Note updated.", "success")
        return redirect(f"/notes/{nid}")
    return render_template("notes/edit.html", note=note)

@app.route("/notes/<int:nid>/delete", methods=["POST"])
@login_required
def delete_note(nid):
    db = get_db()
    db.execute("DELETE FROM notes WHERE id = ? AND user_id = ?",
               (nid, session["user_id"]))
    db.commit()
    flash("Note deleted.", "success")
    return redirect("/dashboard")

# ══════════════════════════════════════════════════════════════════
#  GANTT CHARTS
# ══════════════════════════════════════════════════════════════════

def parse_gantt_csv(file_content):
    """Parse a marketing Gantt CSV and return structured data."""
    reader = csv.reader(io.StringIO(file_content))
    rows = list(reader)
    if not rows:
        return None

    header = rows[0]

    # Find week date columns and Objective/KPI positions
    week_dates = []
    objective_col = None
    kpi_col = None

    for i, col in enumerate(header):
        col_stripped = col.strip()
        if col_stripped == "Objective":
            objective_col = i
        elif col_stripped == "KPI":
            kpi_col = i
            break

    # Week dates are between column 3 and the Objective column
    end_col = objective_col if objective_col else len(header)
    for i in range(3, end_col):
        col_stripped = header[i].strip()
        if not col_stripped:
            continue
        try:
            dt = datetime.strptime(col_stripped, "%d-%b-%y")
            week_dates.append((i, dt.strftime("%Y-%m-%d")))
        except ValueError:
            pass

    if not week_dates:
        return None

    all_week_strs = [wd[1] for wd in week_dates]

    # Parse data rows
    parsed_rows = []
    sort_order = 0
    for row in rows[1:]:
        if not row:
            continue

        action = row[0].strip() if len(row) > 0 else ""
        owner = row[1].strip() if len(row) > 1 else ""
        hours_str = row[2].strip() if len(row) > 2 else ""

        if not action:
            continue

        # Skip summary rows
        if action in ("Total Hours", "Remaining Hours"):
            continue

        hours = 0.0
        try:
            hours = float(hours_str)
        except (ValueError, TypeError):
            pass

        objective = ""
        if objective_col and len(row) > objective_col:
            objective = row[objective_col].strip()
        kpi = ""
        if kpi_col and len(row) > kpi_col:
            kpi = row[kpi_col].strip()

        # Check week cells for content
        week_notes = {}
        first_week = None
        last_week = None
        for col_idx, week_date in week_dates:
            if len(row) > col_idx and row[col_idx].strip():
                week_notes[week_date] = row[col_idx].strip()
                if first_week is None:
                    first_week = week_date
                last_week = week_date

        # Determine row type: category (no owner, no hours, no objective) vs task
        is_task = bool(owner) or hours > 0 or bool(objective)
        row_type = "task" if is_task else "category"

        # For tasks: determine date span
        if row_type == "task":
            start_week = first_week or all_week_strs[0]
            end_week = last_week or all_week_strs[-1]
        else:
            start_week = ""
            end_week = ""

        parsed_rows.append({
            "row_type": row_type,
            "action": action,
            "owner": owner,
            "hours": hours,
            "objective": objective,
            "kpi": kpi,
            "start_week": start_week,
            "end_week": end_week,
            "notes_json": json.dumps(week_notes),
            "sort_order": sort_order,
        })
        sort_order += 1

    return {"weeks": all_week_strs, "rows": parsed_rows}


def parse_gantt_xlsx(file_bytes):
    """Parse a marketing Gantt XLSX and return structured data.

    Supports two layouts:
    1. CSV-like: Action | Owner | Hours | week dates (DD-MMM-YY) | Objective | KPI
    2. Color-based: Title row, month row, then header row with short dates (e.g. "1 Apr"),
       Gantt bars indicated by colored cell backgrounds rather than text content.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Also load with formatting to read cell colors
    wb_fmt = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws_fmt = wb_fmt.active

    # Identify merged cell rows (full-width merges = category rows)
    cat_rows = set()
    for mr in ws.merged_cells.ranges:
        rng = str(mr)
        if rng.startswith("A") and mr.min_row == mr.max_row and (mr.max_col - mr.min_col) >= 4:
            cat_rows.add(mr.min_row)

    # Find the header row — look for a row containing "Owner" in column B
    header_row_idx = None
    for row_idx in range(1, min(ws.max_row + 1, 10)):
        cell_b = ws.cell(row=row_idx, column=2).value
        if cell_b and str(cell_b).strip().lower() == "owner":
            header_row_idx = row_idx
            break

    if header_row_idx is None:
        return None

    header = [str(ws.cell(row=header_row_idx, column=c).value or "").strip()
              for c in range(1, ws.max_column + 1)]

    # Detect layout: find where week columns start and what format they use
    # Check for Hours column and Objective/KPI columns
    hours_col = None
    objective_col = None
    kpi_col = None
    for i, col in enumerate(header):
        col_lower = col.lower()
        if col_lower == "hours":
            hours_col = i
        elif col_lower == "objective":
            objective_col = i
        elif col_lower == "kpi":
            kpi_col = i

    # Determine first week column (after Owner, and optionally Hours)
    first_week_col = 2  # default: column C (index 2)
    if hours_col is not None:
        first_week_col = hours_col + 1

    # Determine the year from the spreadsheet (check row 1 or 2 for year mention)
    inferred_year = None
    for row_idx in range(1, header_row_idx):
        for col_idx in range(1, min(ws.max_column + 1, 20)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                import re as _re
                year_match = _re.search(r"20\d{2}", str(val))
                if year_match:
                    inferred_year = int(year_match.group())
                    break
        if inferred_year:
            break
    if not inferred_year:
        inferred_year = datetime.now().year

    # Parse week date columns
    end_col = objective_col if objective_col else len(header)
    week_dates = []
    for i in range(first_week_col, end_col):
        col_stripped = header[i]
        if not col_stripped:
            continue
        # Try DD-MMM-YY format first (CSV-like layout)
        try:
            dt = datetime.strptime(col_stripped, "%d-%b-%y")
            week_dates.append((i, dt.strftime("%Y-%m-%d")))
            continue
        except ValueError:
            pass
        # Try short format like "1 Apr", "13 May" (color-based layout)
        try:
            dt = datetime.strptime(f"{col_stripped} {inferred_year}", "%d %b %Y")
            week_dates.append((i, dt.strftime("%Y-%m-%d")))
            continue
        except ValueError:
            pass

    if not week_dates:
        return None

    all_week_strs = [wd[1] for wd in week_dates]

    # Colors to ignore when detecting Gantt bars (white / no fill)
    white_fills = {"FFFFFFFF", "00000000"}

    # Parse data rows
    parsed_rows = []
    sort_order = 0
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        action_val = ws.cell(row=row_idx, column=1).value
        if action_val is None:
            continue
        action = str(action_val).strip()
        if not action:
            continue

        # Skip summary / legend rows
        if action.upper() in ("TOTAL HOURS", "REMAINING HOURS", "COLOUR KEY:"):
            continue
        # Skip milestone-like header rows without useful data
        if action.upper() == "KEY MILESTONES":
            continue

        owner = str(ws.cell(row=row_idx, column=2).value or "").strip()

        hours = 0.0
        if hours_col is not None:
            hours_val = ws.cell(row=row_idx, column=hours_col + 1).value
            if hours_val is not None:
                try:
                    hours = float(hours_val)
                except (ValueError, TypeError):
                    pass

        objective = ""
        if objective_col is not None:
            obj_val = ws.cell(row=row_idx, column=objective_col + 1).value
            if obj_val:
                objective = str(obj_val).strip()
        kpi = ""
        if kpi_col is not None:
            kpi_val = ws.cell(row=row_idx, column=kpi_col + 1).value
            if kpi_val:
                kpi = str(kpi_val).strip()

        # Check week cells — look for text content OR colored backgrounds (Gantt bars)
        week_notes = {}
        first_week = None
        last_week = None
        for col_idx, week_date in week_dates:
            cell_val = ws.cell(row=row_idx, column=col_idx + 1).value
            cell_text = str(cell_val).strip() if cell_val else ""

            # Check for colored fill (Gantt bar)
            fmt_cell = ws_fmt.cell(row=row_idx, column=col_idx + 1)
            has_color = False
            if fmt_cell.fill and fmt_cell.fill.fgColor and fmt_cell.fill.fgColor.rgb:
                rgb = str(fmt_cell.fill.fgColor.rgb)
                if rgb not in white_fills:
                    has_color = True

            if cell_text or has_color:
                week_notes[week_date] = cell_text if cell_text else ""
                if first_week is None:
                    first_week = week_date
                last_week = week_date

        # Determine row type
        is_category = row_idx in cat_rows and not owner
        row_type = "category" if is_category else "task"

        if row_type == "task":
            start_week = first_week or all_week_strs[0]
            end_week = last_week or all_week_strs[-1]
        else:
            start_week = ""
            end_week = ""

        parsed_rows.append({
            "row_type": row_type,
            "action": action,
            "owner": owner,
            "hours": hours,
            "objective": objective,
            "kpi": kpi,
            "start_week": start_week,
            "end_week": end_week,
            "notes_json": json.dumps(week_notes),
            "sort_order": sort_order,
        })
        sort_order += 1

    return {"weeks": all_week_strs, "rows": parsed_rows}


@app.route("/gantt")
@login_required
def gantt_list():
    db = get_db()
    groups = db.execute(
        "SELECT * FROM gantt_groups WHERE user_id = ? ORDER BY updated DESC",
        (session["user_id"],)
    ).fetchall()
    ungrouped = db.execute(
        "SELECT * FROM gantt_charts WHERE user_id = ? AND (group_id IS NULL OR group_id = 0) ORDER BY updated DESC",
        (session["user_id"],)
    ).fetchall()
    # Attach charts to each group
    groups_with_charts = []
    for g in groups:
        charts = db.execute(
            "SELECT * FROM gantt_charts WHERE group_id = ? ORDER BY updated DESC",
            (g["id"],)
        ).fetchall()
        groups_with_charts.append({"group": g, "charts": charts})
    return render_template("gantt/list.html", groups=groups_with_charts, ungrouped=ungrouped)


@app.route("/gantt/group/new", methods=["GET", "POST"])
@login_required
def gantt_group_new():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("Please enter a group name.", "error")
            return redirect("/gantt/group/new")
        db = get_db()
        now = datetime.now(timezone.utc).isoformat()
        cursor = db.execute(
            "INSERT INTO gantt_groups (user_id, name, created, updated) VALUES (?, ?, ?, ?)",
            (session["user_id"], name, now, now)
        )
        db.commit()
        flash(f"Group '{name}' created.", "success")
        return redirect(f"/gantt/group/{cursor.lastrowid}")
    return render_template("gantt/group_new.html")


@app.route("/gantt/group/<int:group_id>")
@login_required
def gantt_group_view(group_id):
    db = get_db()
    group = db.execute(
        "SELECT * FROM gantt_groups WHERE id = ? AND user_id = ?",
        (group_id, session["user_id"])
    ).fetchone()
    if not group:
        abort(404)
    charts = db.execute(
        "SELECT * FROM gantt_charts WHERE group_id = ? ORDER BY id",
        (group_id,)
    ).fetchall()
    # If there are charts, show the first one by default
    if charts:
        return redirect(f"/gantt/{charts[0]['id']}")
    return render_template("gantt/group_view.html", group=group, charts=charts)


@app.route("/gantt/group/<int:group_id>/delete", methods=["POST"])
@login_required
def gantt_group_delete(group_id):
    db = get_db()
    # Ungroup charts (don't delete them)
    db.execute("UPDATE gantt_charts SET group_id = NULL WHERE group_id = ?", (group_id,))
    db.execute("DELETE FROM gantt_groups WHERE id = ? AND user_id = ?", (group_id, session["user_id"]))
    db.commit()
    flash("Group deleted. Charts have been ungrouped.", "success")
    return redirect("/gantt")


@app.route("/gantt/import", methods=["GET", "POST"])
@login_required
def gantt_import():
    if request.method == "POST":
        file = request.files.get("csv_file")
        chart_name = request.form.get("name", "").strip()
        group_id = request.form.get("group_id", "").strip()
        brand_color = request.form.get("brand_color", "#2a5a8a").strip()

        if not file or not file.filename:
            flash("Please select a file.", "error")
            return redirect("/gantt/import")

        filename_lower = file.filename.lower()
        if not filename_lower.endswith((".csv", ".xlsx")):
            flash("Please upload a CSV or XLSX file.", "error")
            return redirect("/gantt/import")

        if not chart_name:
            chart_name = file.filename.rsplit(".", 1)[0]

        raw = file.read()

        if filename_lower.endswith(".xlsx"):
            parsed = parse_gantt_xlsx(raw)
        else:
            try:
                content = raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                content = raw.decode("latin-1")
            parsed = parse_gantt_csv(content)
        if not parsed or not parsed["rows"]:
            flash("Could not parse file. Check the format.", "error")
            return redirect("/gantt/import")

        db = get_db()
        now = datetime.now(timezone.utc).isoformat()

        gid = int(group_id) if group_id else None

        cursor = db.execute(
            "INSERT INTO gantt_charts (user_id, name, weeks_json, group_id, brand_color, created, updated) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (session["user_id"], chart_name, json.dumps(parsed["weeks"]), gid, brand_color, now, now)
        )
        chart_id = cursor.lastrowid

        for row in parsed["rows"]:
            db.execute(
                """INSERT INTO gantt_rows
                   (chart_id, row_type, action, owner, hours, objective, kpi,
                    status, start_week, end_week, notes_json, sort_order)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (chart_id, row["row_type"], row["action"], row["owner"],
                 row["hours"], row["objective"], row["kpi"],
                 "not_started", row["start_week"], row["end_week"],
                 row["notes_json"], row["sort_order"])
            )

        db.commit()
        flash(f"Imported '{chart_name}' with {len(parsed['rows'])} rows.", "success")
        return redirect(f"/gantt/{chart_id}", code=303)

    db = get_db()
    groups = db.execute(
        "SELECT * FROM gantt_groups WHERE user_id = ? ORDER BY name",
        (session["user_id"],)
    ).fetchall()
    group_id = request.args.get("group_id", "")
    return render_template("gantt/import.html", groups=groups, selected_group_id=group_id)


@app.route("/gantt/<int:chart_id>")
@login_required
def gantt_view(chart_id):
    db = get_db()
    chart = db.execute(
        "SELECT * FROM gantt_charts WHERE id = ? AND user_id = ?",
        (chart_id, session["user_id"])
    ).fetchone()
    if not chart:
        abort(404)

    rows = db.execute(
        "SELECT * FROM gantt_rows WHERE chart_id = ? ORDER BY sort_order",
        (chart_id,)
    ).fetchall()

    weeks = json.loads(chart["weeks_json"])

    # Group weeks by month for header
    month_groups = []
    current_month = None
    today_str = datetime.now().strftime("%Y-%m-%d")
    scroll_to_week_index = 0
    for idx, w in enumerate(weeks):
        dt = datetime.strptime(w, "%Y-%m-%d")
        month_label = dt.strftime("%b %Y")
        if month_label != current_month:
            month_groups.append({"label": month_label, "count": 1, "start_index": idx})
            current_month = month_label
        else:
            month_groups[-1]["count"] += 1
        # Find the week closest to today for auto-scroll
        if w <= today_str:
            scroll_to_week_index = idx

    # Find the week that contains today (for column highlight)
    today_week = ""
    for i, w in enumerate(weeks):
        if w <= today_str:
            today_week = w
        else:
            break

    # Format week labels
    week_labels = []
    for w in weeks:
        dt = datetime.strptime(w, "%Y-%m-%d")
        week_labels.append({"date": w, "label": dt.strftime("%d %b")})

    # Process rows for template
    processed_rows = []
    for row in rows:
        notes = json.loads(row["notes_json"]) if row["notes_json"] else {}
        processed_rows.append({
            "id": row["id"],
            "row_type": row["row_type"],
            "action": row["action"],
            "owner": row["owner"],
            "hours": row["hours"],
            "objective": row["objective"],
            "kpi": row["kpi"],
            "status": row["status"],
            "start_week": row["start_week"],
            "end_week": row["end_week"],
            "notes": notes,
        })

    # Stats
    tasks = [r for r in processed_rows if r["row_type"] == "task"]
    total_tasks = len(tasks)
    complete_count = sum(1 for t in tasks if t["status"] == "complete")
    total_hours = sum(t["hours"] for t in tasks)

    # Sibling charts in the same group (for tabs)
    group = None
    sibling_charts = []
    if chart["group_id"]:
        group = db.execute(
            "SELECT * FROM gantt_groups WHERE id = ?", (chart["group_id"],)
        ).fetchone()
        sibling_charts = db.execute(
            "SELECT id, name, logo_path, brand_color FROM gantt_charts WHERE group_id = ? ORDER BY id",
            (chart["group_id"],)
        ).fetchall()

    return render_template("gantt/view.html",
                           chart=chart, rows=processed_rows,
                           weeks=weeks, week_labels=week_labels,
                           month_groups=month_groups,
                           total_tasks=total_tasks, complete_count=complete_count,
                           total_hours=total_hours,
                           group=group, sibling_charts=sibling_charts,
                           scroll_to_week_index=scroll_to_week_index,
                           today=today_str,
                           today_week=today_week,
                           today_formatted=datetime.now().strftime("%d %b %Y"))


@app.route("/gantt/<int:chart_id>/settings", methods=["GET", "POST"])
@login_required
def gantt_settings(chart_id):
    db = get_db()
    chart = db.execute(
        "SELECT * FROM gantt_charts WHERE id = ? AND user_id = ?",
        (chart_id, session["user_id"])
    ).fetchone()
    if not chart:
        abort(404)

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        brand_color = request.form.get("brand_color", "#2a5a8a").strip()
        group_id = request.form.get("group_id", "").strip()

        if name:
            db.execute("UPDATE gantt_charts SET name = ? WHERE id = ?", (name, chart_id))

        db.execute("UPDATE gantt_charts SET brand_color = ? WHERE id = ?", (brand_color, chart_id))

        gid = int(group_id) if group_id else None
        db.execute("UPDATE gantt_charts SET group_id = ? WHERE id = ?", (gid, chart_id))

        # Handle logo upload
        logo_file = request.files.get("logo")
        if logo_file and logo_file.filename:
            ext = logo_file.filename.rsplit(".", 1)[-1].lower()
            if ext in ("png", "jpg", "jpeg", "gif", "svg", "webp"):
                logo_dir = UPLOAD_DIR / "logos"
                logo_dir.mkdir(exist_ok=True)
                fname = f"chart_{chart_id}_{secure_filename(logo_file.filename)}"
                logo_file.save(str(logo_dir / fname))
                db.execute("UPDATE gantt_charts SET logo_path = ? WHERE id = ?", (f"logos/{fname}", chart_id))

        db.execute("UPDATE gantt_charts SET updated = ? WHERE id = ?",
                    (datetime.now(timezone.utc).isoformat(), chart_id))
        db.commit()
        flash("Chart settings updated.", "success")
        return redirect(f"/gantt/{chart_id}")

    groups = db.execute(
        "SELECT * FROM gantt_groups WHERE user_id = ? ORDER BY name",
        (session["user_id"],)
    ).fetchall()
    return render_template("gantt/settings.html", chart=chart, groups=groups)


@app.route("/gantt/<int:chart_id>/update-task", methods=["POST"])
@login_required
def gantt_update_task(chart_id):
    db = get_db()
    chart = db.execute(
        "SELECT id FROM gantt_charts WHERE id = ? AND user_id = ?",
        (chart_id, session["user_id"])
    ).fetchone()
    if not chart:
        return jsonify({"error": "Not found"}), 404

    data = request.get_json() if request.is_json else {}
    row_id = data.get("row_id")
    status = data.get("status")

    if not row_id or status not in ("complete", "on_track", "behind", "not_started"):
        return jsonify({"error": "Invalid data"}), 400

    db.execute(
        "UPDATE gantt_rows SET status = ? WHERE id = ? AND chart_id = ?",
        (status, row_id, chart_id)
    )
    db.execute(
        "UPDATE gantt_charts SET updated = ? WHERE id = ?",
        (datetime.now(timezone.utc).isoformat(), chart_id)
    )
    db.commit()
    return jsonify({"ok": True})


@app.route("/gantt/<int:chart_id>/delete", methods=["POST"])
@login_required
def gantt_delete(chart_id):
    db = get_db()
    chart = db.execute(
        "SELECT group_id FROM gantt_charts WHERE id = ? AND user_id = ?",
        (chart_id, session["user_id"])
    ).fetchone()
    db.execute(
        "DELETE FROM gantt_charts WHERE id = ? AND user_id = ?",
        (chart_id, session["user_id"])
    )
    db.commit()
    flash("Chart deleted.", "success")
    if chart and chart["group_id"]:
        return redirect(f"/gantt/group/{chart['group_id']}")
    return redirect("/gantt")


# ══════════════════════════════════════════════════════════════════
#  PLATFORM ADMIN — /admin/
# ══════════════════════════════════════════════════════════════════

@app.route("/admin")
@platform_admin_required
def admin_dashboard():
    db = get_db()
    tab = request.args.get("tab", "users")
    stats = {
        "total_users": db.execute("SELECT COUNT(*) FROM users").fetchone()[0],
        "new_users_7d": db.execute(
            "SELECT COUNT(*) FROM users WHERE created >= datetime('now', '-7 days')"
        ).fetchone()[0],
    }
    users = db.execute("""
        SELECT u.*
        FROM users u ORDER BY u.created DESC
    """).fetchall()

    activity = []
    for u in db.execute("SELECT id AS user_id, name AS user_name, created FROM users ORDER BY created DESC LIMIT 30").fetchall():
        activity.append({"type": "signup", "user_name": u["user_name"], "user_id": u["user_id"],
                         "detail": None, "created": u["created"]})
    activity.sort(key=lambda x: x["created"], reverse=True)
    activity = activity[:50]

    return render_template("admin.html", stats=stats, users=users, activity=activity, tab=tab)

@app.route("/admin/users/<int:uid>/toggle-admin", methods=["POST"])
@platform_admin_required
def admin_toggle_platform_admin(uid):
    if uid == session["user_id"]:
        flash("Cannot change your own admin status.", "error")
        return redirect("/admin?tab=users")
    db = get_db()
    user = get_user_by_id(uid)
    if not user:
        flash("User not found.", "error")
        return redirect("/admin?tab=users")
    db.execute("UPDATE users SET is_admin = ? WHERE id = ?", (0 if user["is_admin"] else 1, uid))
    db.commit()
    flash(f"{'Removed' if user['is_admin'] else 'Granted'} admin for {user['name']}.", "success")
    return redirect("/admin?tab=users")

@app.route("/admin/users/<int:uid>/delete", methods=["POST"])
@platform_admin_required
def admin_delete_user(uid):
    if uid == session["user_id"]:
        flash("Cannot delete yourself.", "error")
        return redirect("/admin?tab=users")
    db = get_db()
    db.execute("DELETE FROM users WHERE id = ?", (uid,))
    db.commit()
    flash("User deleted.", "success")
    return redirect("/admin?tab=users")

@app.route("/admin/users/<int:uid>/impersonate", methods=["POST"])
@platform_admin_required
def admin_impersonate(uid):
    user = get_user_by_id(uid)
    if not user:
        flash("User not found.", "error")
        return redirect("/admin?tab=users")
    session["impersonator_id"] = session["user_id"]
    session["user_id"] = user["id"]
    session["user_name"] = user["name"]
    flash(f"Now viewing as {user['name']}. Use the banner to switch back.", "success")
    return redirect("/dashboard")

@app.route("/admin/stop-impersonating", methods=["POST"])
@login_required
def admin_stop_impersonating():
    real_id = session.pop("impersonator_id", None)
    if real_id:
        user = get_user_by_id(real_id)
        if user:
            session["user_id"] = user["id"]
            session["user_name"] = user["name"]
            flash("Switched back to your admin account.", "success")
            return redirect("/admin")
    return redirect("/dashboard")

# ── Legal Pages ───────────────────────────────────────────────────

@app.route("/terms")
def terms_page():
    return render_template("legal.html", title="Terms of Service",
                           content="These terms of service govern your use of this platform. By using this platform, you agree to these terms. This is a placeholder — update with your actual terms before launch.")

@app.route("/privacy")
def privacy_page():
    return render_template("legal.html", title="Privacy Policy",
                           content="We collect your name, email, and content you create. We do not sell your data to third parties. Cookies are used for session management only. This is a placeholder — update with your actual privacy policy before launch.")

# ── Error Handlers ────────────────────────────────────────────────

@app.errorhandler(404)
def page_not_found(e):
    return render_template("errors/404.html"), 404

@app.errorhandler(500)
def internal_error(e):
    return render_template("errors/500.html"), 500

@app.errorhandler(429)
def rate_limited(e):
    return render_template("errors/429.html"), 429

# ── Boot ──────────────────────────────────────────────────────────

init_db()

if __name__ == "__main__":
    app.run(debug=os.environ.get("FLASK_DEBUG", "0") == "1", host="0.0.0.0", port=DEFAULT_PORT)
